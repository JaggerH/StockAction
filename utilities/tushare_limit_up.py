import os
import tushare as ts
import pandas as pd
from datetime import datetime

# Mapping English column names to Chinese
columns_mapping = {
    # 'trade_date': '交易日期',
    'ts_code': '股票代码',
    'name': '股票名称',
    'industry': '所属行业',
    'close': '收盘价',
    'pct_chg': '涨跌幅',
    'amount': '成交额(亿)',
    'circ_mv': '流通市值(亿)',
    'total_mv': '总市值(亿)',
    'pb': 'PB',
    'pe': 'PE',
    'turnover_rate': '换手率',
    'turnover_rate_f': '流通股换手率',
    # 'limit_amount': '板上成交金额(亿)',
    'fd_amount': '封单金额(亿)',
    'first_time': '首次封板时间',
    'last_time': '最后封板时间',
    'open_times': '炸板次数',
    'up_stat': '涨停统计',
    'limit_times': '连板数',
}

def prepare_df(df):
    # Order the DataFrame by 'limit_times' in descending order
    limit_up_df_copy = df.sort_values(by='limit_times', ascending=False)
    # 预处理单位
    limit_up_df_copy['amount'] = limit_up_df_copy['amount'] * 1000 # 成交额 原单位应是千分位
    limit_up_df_copy['circ_mv'] = limit_up_df_copy['circ_mv'] * 10000 # 流值 原单位应是万
    limit_up_df_copy['total_mv'] = limit_up_df_copy['total_mv'] * 10000 # 总值 原单位应是万

    # 将金额单位全部换成亿
    limit_up_df_copy['amount'] = limit_up_df_copy['amount'] / 10000 / 10000
    limit_up_df_copy['limit_amount'] = limit_up_df_copy['limit_amount'] / 10000 / 10000
    limit_up_df_copy['circ_mv'] = limit_up_df_copy['circ_mv'] / 10000 / 10000
    limit_up_df_copy['total_mv'] = limit_up_df_copy['total_mv'] / 10000 / 10000
    limit_up_df_copy['fd_amount'] = limit_up_df_copy['fd_amount'] / 10000 / 10000

    columns_sequence = columns_mapping.keys()
    limit_up_df_copy = limit_up_df_copy.reindex(columns=columns_sequence)
    # Rename the columns using the mapping
    return limit_up_df_copy

def build_xlsx(df, path, ths_concept=None):
    # 定义一个函数将 name 转换为超链接格式
    def make_name_hyperlink(row):
        if row["name"].startswith("=HYPERLINK"):
            return row["name"]
        symbol, exchange = row['ts_code'].split(".")
        url = f'https://quote.eastmoney.com/{exchange.lower() + symbol}.html'
        return f'=HYPERLINK("{url}", "{row["name"]}")'
    
    def make_industry_hyperlink(row):
        if row["industry"].startswith("=HYPERLINK"):
            return row["industry"]
        symbol, exchange = row['ts_code'].split(".")
        url = f'https://emweb.securities.eastmoney.com/pc_hsf10/pages/index.html?type=web&code={exchange + symbol}&color=b#/hxtc'
        return f'=HYPERLINK("{url}", "{row["industry"]}")'

    def make_industry_by_concept(row):
        code = row['ts_code']
        concept_df = ths_concept[ths_concept['code'] == code]
        # 按绝对值降序排列
        concept_df = concept_df.reindex(concept_df.sort_values(by='pct_change', key=abs, ascending=False).index)
        concept_df = concept_df[:3] # 只取前三项，多了显示麻烦
        return '\n'.join(concept_df.apply(lambda row: f"{row['concept']}({row['pct_change']: .2f}%)", axis=1))

    # 应用函数到 name 列
    # df['name'] = df.apply(make_name_hyperlink, axis=1)

    # 由于更新了详细的相关概念，原有的超链接关闭
    # df['industry'] = df.apply(make_industry_hyperlink, axis=1)
    if ths_concept is not None:
        df['industry'] = df.apply(make_industry_by_concept, axis=1)

    df = df.rename(columns=columns_mapping)
    df.to_excel(path, index=False)

def beautify_xlsx(path):
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    
    # Create a new Excel file with formatted values
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # 冻结行首
    ws.freeze_panes = 'A2'

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                if cell.value is not None:
                    cell.value = round(cell.value, 2)
                    cell.number_format = '0.00'

    for idx, col in enumerate(ws.columns, 1):
        max_length = 5
        column = get_column_letter(idx)  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if column in ["C"]: 
                    cell.alignment = Alignment(wrap_text=True)
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
        # 特殊列，含中文字符、表达式，宽度自动适配
        if column in ["B", "C"]: 
            if column in ["C"]: 
                ws.column_dimensions[column].width = 20
            ws.column_dimensions[get_column_letter(idx)].auto_size = True
        else:
            adjusted_width = max_length + 4
            ws.column_dimensions[column].width = adjusted_width

    # 表头自动换行且居中
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Save the workbook
    wb.save(path)
    return path
def generate_limit_up_excel(df, cal_date, path=None, ths_concepts=None) -> str:
    """
        在运行该函数前需要运行 generate_limit_up_df 获取数据
        df, cal_date = generate_limit_up_df()
    """
    template_path = '/tmp/涨停分析-%s.xlsx' if os.getenv("ENV") == "product" else './涨停分析-%s.xlsx'
    excel_file_path = template_path % cal_date if path is None else path
    build_xlsx(df, excel_file_path, ths_concepts)
    beautify_xlsx(excel_file_path)

    return excel_file_path

def sendLimitUpEmail():
    import logging
    import datetime

    from utilities.set_env import set_env
    from utilities.azure_cosmos import init_container
    from utilities.email_client import EmailClient
    import azure.cosmos.exceptions as exceptions

    # for limit_up_trigger_validate
    import uuid
    utc_timestamp = datetime.datetime.utcnow().replace(
        tzinfo=datetime.timezone.utc).isoformat()
    # ------------------

    set_env()
    container = init_container()
    instance = TushareLimitUp()
    cal_date = instance.getLatestTradeData()
    try:
        # 如果能正常读取limit_up_identifier，那是已经发送过通知了
        record = container.read_item(item=cal_date, partition_key="limit_up_identifier")
        # for limit_up_trigger_validate
        container.create_item(body={
            "id": str(uuid.uuid4()),
            "partitionKey": "limit_up_trigger_validate",
            'type': 'skip',
            "log": 'limit up has created, skip.',
            "created_at": utc_timestamp
        })
        logging.info('limit up has created, skip.')
    except exceptions.CosmosResourceNotFoundError:
        try:
            df = instance.generate_limit_up_df()
            ths_concepts_exist_df = instance.pro.ths_member(**{ "ts_code": cal_date, "limit": "10", }, fields=[ "ts_code", "code", "name" ])
            if instance.limit_up_df.empty or instance.daily_basic_df.empty or ths_concepts_exist_df.empty:
                logging.info('tushare has not update')
                # for limit_up_trigger_validate
                container.create_item(body={
                    "id": str(uuid.uuid4()),
                    "partitionKey": "limit_up_trigger_validate",
                    'type': 'skip',
                    "log": 'tushare has not update',
                    "created_at": utc_timestamp
                })
            else:
                ths_concepts = instance.readThsConcept(cal_date)
                path = generate_limit_up_excel(df, cal_date, ths_concepts=ths_concepts)
                logging.info('limit up file created at %s' % path)
                # for limit_up_trigger_validate
                container.create_item(body={
                    "id": str(uuid.uuid4()),
                    "partitionKey": "limit_up_trigger_validate",
                    'type': 'success',
                    "log": 'limit up file created at %s' % path,
                    "created_at": utc_timestamp
                })
                
                utc_timestamp = datetime.datetime.utcnow().replace(
                    tzinfo=datetime.timezone.utc).isoformat()

                email = EmailClient(
                    os.getenv("SENDER_EMAIL"),
                    os.getenv("SENDER_EMAIL_PASS"),
                    os.getenv("RECEIVERS_EMAIL")
                )
                email.setSubject("%s涨停列表" % cal_date)
                email.addContent("请查阅附件。\n\n")
                email.addContent("顺颂商祺")
                email.addAttachment(path)
                email.sendMail()
                os.remove(path)
                logging.info('email has sent')

                has_send_notify = {
                    "id": instance.cal_date,
                    "partitionKey": "limit_up_identifier",
                    "is_send_notify": True,
                    "created_at": utc_timestamp
                }
                container.create_item(body=has_send_notify)
        except Exception as e:
            logging.error("An error occurred", exc_info=True)

def deleteEmailSentFlag():
    from utilities.azure_cosmos import init_container
    import azure.cosmos.exceptions as exceptions

    container = init_container()
    instance = TushareLimitUp()
    cal_date = instance.getLatestTradeData()
    try:
        container.delete_item(item=cal_date, partition_key="limit_up_identifier")
        print('limit_up_identifier %s has_send_notify flag is deleted' % cal_date)
    except exceptions.CosmosResourceNotFoundError:
        return False

class TushareLimitUp:
    def __init__(self) -> None:
        token = os.getenv('TUSHARE_API_KEY')
        self.pro = ts.pro_api(token)
        self.isReadTushare = False
        self.isReadThsConcept = False

    # 获取交易日历
    def getLatestTradeData(self, specific_date=None) -> str:
        """
            如果今天不是交易日期，将返回最近交易日
        """
        today_str = specific_date if specific_date is not None else datetime.now().strftime("%Y%m%d")
        df = self.pro.trade_cal(**{ "exchange": "SSE", "cal_date": today_str }, fields=[ "exchange", "cal_date", "is_open", "pretrade_date" ])
        row = df.loc[0]
        return row['pretrade_date'] if row['is_open'] == 0 else row['cal_date']

    def readTushareData(self, cal_date) -> None:
        # 拉取数据
        self.limit_up_df = self.pro.limit_list_d(**{ "trade_date": cal_date, "limit_type": "U", }, fields=[ "trade_date", "ts_code", "limit_amount", "fd_amount", "first_time", "last_time", "open_times", "up_stat", "limit_times", "limit" ])
        self.daily_df = self.pro.daily(**{"trade_date": cal_date}, fields=[ "ts_code", "trade_date", "open", "high", "low", "close", "pre_close", "change", "pct_chg", "vol", "amount" ])

        self.stock_basic_df = self.pro.stock_basic(**{"list_status": 'L'}, fields=['ts_code','name','industry'])
        self.daily_basic_df = self.pro.daily_basic(**{"trade_date": cal_date}, fields=['ts_code','trade_date','circ_mv','total_mv','turnover_rate','turnover_rate_f','pe','pb'])

        self.isReadTushare = True

    def readThsConcept(self, cal_date) -> pd.DataFrame:
        # 以下是同花顺概念
        ## 取同花顺所有概念 exchange -> A-a股，type -> N-概念指数
        self.ths_index = self.pro.ths_index(**{ "exchange": "A", "type": "N" }, fields=[ "ts_code", "name", "count", "exchange", "list_date", "type" ])
        self.ths_index = self.ths_index.rename(columns={'name': 'concept'})
        ## 取同花顺概念当日涨幅
        self.ths_daily = self.pro.ths_daily(**{ "trade_date": cal_date }, fields=[ "ts_code", "trade_date", "close", "open", "high", "low", "pre_close", "avg_price", "change", "pct_change", "vol", "turnover_rate" ])
        self.ths_index = pd.merge(self.ths_index, self.ths_daily, on='ts_code', how='inner')

        limit_up_df = self.getLimitUpData(cal_date)
        self.ths_members = [ self.pro.ths_member(**{ "code": code }, fields=[ "ts_code", "code", "name" ]) for code in list(limit_up_df['ts_code']) ]
        self.ths_members = pd.concat(self.ths_members, ignore_index=True)

        self.ths_concept = pd.merge(self.ths_members, self.ths_index, on='ts_code', how='inner')
        self.isReadThsConcept = True
        return self.ths_concept

    # 涨停数据
    def getLimitUpData(self, cal_date) -> pd.DataFrame:
        """
            Tushare生成imit_list_d数据时间段不固定，但是limit_list_d出来之前，会缺失涨停数据
            如果检测到数据不为空，会写入数据库
        """
        # 读取过就不再重复读取，要强制就在执行readTushareData之前把isReadTushare设置为False
        if not self.isReadTushare:
            self.readTushareData(cal_date)
            
        self.tenpct_df = self.daily_df[self.daily_df['pct_chg'] >= 10]

        # 合并tenpct_df和limit_up_df
        union_tscode = pd.concat([self.tenpct_df['ts_code'], self.limit_up_df['ts_code']]).drop_duplicates()
        union_tscode_df = self.daily_df[self.daily_df['ts_code'].isin(union_tscode)]
        merged_df = pd.merge(union_tscode_df, self.stock_basic_df, on=['ts_code'], how='left')
        merged_df = pd.merge(merged_df, self.daily_basic_df, on=['ts_code'], how='left')
        merged_df = pd.merge(merged_df, self.limit_up_df, on=['ts_code'], how='outer')

        return merged_df

    def generate_limit_up_df(self, specific_date=None) -> pd.DataFrame:
        # 如果有指定日期，使用指定日期，否则使用最近的交易日
        self.cal_date = self.getLatestTradeData(specific_date)

        df = self.getLimitUpData(self.cal_date)
        df = prepare_df(df)
        return df